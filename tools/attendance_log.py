#!/usr/bin/env python3
"""Attendance logger for the absensi-rfid project.

Reads UID lines from the ESP8266 USB serial port and writes them to a CSV
file with a timestamp. Also runs a tiny local web server so you can:
  - view attendance live (auto-refresh) in the browser
  - add/edit names in the browser (saved to roster.csv)
  - export to Excel (downloads the CSV)

Usage:
    python attendance_log.py            # listen + open browser
    python attendance_log.py --list     # print roster + names, then exit
    python attendance_log.py --no-browser  # listen without opening browser
"""

import argparse
import csv
import datetime
import glob
import http.server
import io
import os
import subprocess
import threading
import time
import urllib.parse
import webbrowser

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import serial

# --- CONFIG -----------------------------------------------------------------
PORT = "COM3"                 # ESP8266 USB serial port
BAUD = 115200
CSV_FILE = "attendance.csv"   # attendance log
ROSTER_FILE = "roster.csv"    # UID -> name (edit in browser or Notepad)
HTTP_PORT = 8080              # local web dashboard
# ponytail: dedup mati secara default buat testing (tap berulang ke-log semua).
# Nyalain (True) buat mode produksi biar 1 orang = 1 check-in per hari.
DEDUP = False
# -----------------------------------------------------------------------------

# UID -> name mapping. Loaded from ROSTER_FILE on startup; editable in browser.
NAME_MAP = {}

# Everyone who is expected to attend. Used to report absentees.
ROSTER = ["Budi", "Rehan"]


def register_from_csv():
    try:
        with open(ROSTER_FILE) as f:
            for uid, name in csv.reader(f):
                if uid:
                    NAME_MAP[uid.strip()] = name.strip()
    except FileNotFoundError:
        pass


def save_roster(uid, name):
    NAME_MAP[uid] = name
    rows = sorted(NAME_MAP.items())
    with open(ROSTER_FILE, "w", newline="") as f:
        csv.writer(f).writerows(rows)


def write_log(uid, name):
    today = datetime.date.today().isoformat()
    if DEDUP and glob.glob(CSV_FILE) and any(row[0].startswith(today) and row[1] == uid
                                  for row in csv.reader(open(CSV_FILE))):
        return  # ponytail: skip duplicate check-in per person per day
    stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new = not glob.glob(CSV_FILE)
    with open(CSV_FILE, "a", newline="") as f:
        w = csv.writer(f)
        if new:
            w.writerow(["time", "uid", "name"])
        w.writerow([stamp, uid, name])
    print(f"[LOG] {stamp}  {name or uid}")


def read_csv():
    rows = []
    if glob.glob(CSV_FILE):
        with open(CSV_FILE) as f:
            rd = csv.reader(f)
            next(rd, None)  # skip header row
            for time_, uid, name in rd:
                rows.append((time_, uid, name))
    return rows


def page_attendance():
    rows = read_csv()
    today = datetime.date.today().isoformat()
    body = "".join(
        f"<tr><td>{t}</td><td><code>{u}</code></td>"
        f"<td>{n} <button class='pen' onclick='rename(\"{u}\",\"{n}\")'>✏️</button>"
        f"<button class='pen' onclick='del({i})'>🗑️</button></td></tr>"
        for i, (t, u, n) in enumerate(rows)
    )
    table = f"<h2>Absensi {today}</h2>" \
            f"<table><tr><th>Jam</th><th>UID</th><th>Nama</th></tr>{body}</table>"
    if not rows:
        table += "<p class='empty'>Belum ada absensi hari ini.</p>"
    nav = "<div class='nav'><a href='/export'>Export Excel</a></div>"
    return "<!DOCTYPE html><html><head><meta charset='utf-8'>" \
           "<meta http-equiv='refresh' content='5'>" \
           "<title>Absensi RFID</title><style>" \
           "body{font-family:Segoe UI,Arial;max-width:640px;margin:30px auto;padding:0 16px;background:#f5f7fa;color:#1f2937}" \
           "h2{color:#2563eb}table{width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.08)}" \
           "th{background:#2563eb;color:#fff;text-align:left;padding:12px}td{padding:12px;border-bottom:1px solid #eee}" \
           "tr:hover td{background:#eff6ff}.empty{color:#6b7280;margin-top:16px}" \
           ".nav{margin:16px 0}.nav a{display:inline-block;background:#2563eb;color:#fff;text-decoration:none;padding:10px 16px;border-radius:8px}" \
           ".pen{border:none;background:none;cursor:pointer;font-size:16px;padding:0 4px}" \
           "</style></head><body>" + nav + table + \
           "<script>async function rename(uid,old){var n=prompt('Nama untuk '+uid,old);" \
           "if(n===null||!n.trim())return;await fetch('/rename',{method:'POST'," \
           "headers:{'Content-Type':'application/x-www-form-urlencoded'}," \
           "body:'uid='+encodeURIComponent(uid)+'&name='+encodeURIComponent(n.trim())});location.reload();}" \
           "async function del(i){if(!confirm('Hapus baris ini?'))return;" \
           "await fetch('/delete',{method:'POST',headers:{'Content-Type':'application/x-www-form-urlencoded'}," \
           "body:'i='+i});location.reload();}</script>" \
           "</body></html>"


def export_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Absensi"
    ws.append(["Jam", "UID", "Nama"])
    for t, u, n in read_csv():
        ws.append([t, u, n])
    for col, w in zip("ABC", (22, 14, 16)):
        ws.column_dimensions[col].width = w
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2563EB")
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class Handler(http.server.BaseHTTPRequestHandler):
    def _send(self, code, body, ctype="text/html; charset=utf-8", extra=None):
        data = body.encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        for k, v in (extra or {}).items():
            self.send_header(k, v)
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        if self.path.split("?")[0] == "/export":
            data = export_xlsx()
            self.send_response(200)
            self.send_header("Content-Type",
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="attendance.xlsx"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
        else:
            self._send(200, page_attendance())

    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(length).decode("utf-8")
        params = urllib.parse.parse_qs(body)
        if self.path.split("?")[0] == "/rename":
            uid = params.get("uid", [""])[0].strip()
            name = params.get("name", [""])[0].strip()
            if uid:
                NAME_MAP[uid] = name
                with open(ROSTER_FILE, "w", newline="") as f:
                    csv.writer(f).writerows(sorted(NAME_MAP.items()))
            self.send_response(303)
            self.send_header("Location", "/")
            self.end_headers()
        elif self.path.split("?")[0] == "/delete":
            try:
                i = int(params.get("i", ["-1"])[0])
                rows = read_csv()
                if 0 <= i < len(rows):
                    rows.pop(i)
                with open(CSV_FILE, "w", newline="") as f:
                    w = csv.writer(f)
                    w.writerow(["time", "uid", "name"])
                    w.writerows(rows)
            except (ValueError, IndexError):
                pass
            self.send_response(303)
            self.send_header("Location", "/")
            self.end_headers()
        else:
            self.send_response(404)
            self.end_headers()

    def log_message(self, *a):
        pass


def _open_browser(url):
    chrome = [os.path.join(p, "Google", "Chrome", "Application", "chrome.exe")
              for p in (os.environ.get("ProgramFiles", ""),
                        os.environ.get("ProgramFiles(x86)", ""))]
    for path in chrome:
        if os.path.exists(path):
            subprocess.Popen([path, url])
            return
    webbrowser.open(url)


def serve():
    server = http.server.ThreadingHTTPServer(("127.0.0.1", HTTP_PORT), Handler)
    threading.Thread(target=server.serve_forever, daemon=True).start()
    url = f"http://127.0.0.1:{HTTP_PORT}"
    _open_browser(url)
    print(f"Dashboard: {url}")


def main():
    register_from_csv()
    ap = argparse.ArgumentParser()
    ap.add_argument("--list", action="store_true")
    ap.add_argument("--no-browser", action="store_true")
    args = ap.parse_args()

    if args.list:
        print("Roster:", ", ".join(ROSTER))
        print("Mapped names:")
        for uid, name in NAME_MAP.items():
            print(f"  {uid}  ->  {name}")
        return

    if not args.no_browser:
        serve()

    while True:
        try:
            s = serial.Serial(PORT, BAUD, timeout=0.5)
            break
        except Exception as e:
            print(f"{PORT} belum siap: {e}. Coba lagi dalam 3 detik...")
            time.sleep(3)
    s.setDTR(False)
    s.setRTS(False)
    s.reset_input_buffer()
    print(f"Listening on {PORT} -> {CSV_FILE}  (Ctrl+C to stop)")

    try:
        while True:
            line = s.readline().decode(errors="replace").strip()
            if line.upper().startswith("UID:"):
                uid = line.split(":", 1)[1].strip()
                name = NAME_MAP.get(uid, "UNKNOWN")
                write_log(uid, name)
    except KeyboardInterrupt:
        print("\nStopped.")
        present = {name for _, _, name in read_csv()}
        absent = [r for r in ROSTER if r not in present]
        print("Absent:", ", ".join(absent) if absent else "everyone present")
    finally:
        s.close()


if __name__ == "__main__":
    main()